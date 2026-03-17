from __future__ import unicode_literals, absolute_import

from collections import namedtuple
from itertools import chain
import csv
import re
import datetime
from dateutil.tz import tzlocal
from tempfile import NamedTemporaryFile

from django.http import HttpResponse
try:
    # Django < 4.0
    from django.utils.encoding import force_text
except ImportError:
    # Django >= 4.0
    from django.utils.encoding import force_str as force_text
from django.template.loader import render_to_string
from django.utils import timezone

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from six import BytesIO, text_type

from .introspection import get_model_from_path_string


DisplayField = namedtuple("DisplayField", "path field")


def generate_filename(title, ends_with):
    title = title.split('.')[0]
    title.replace(' ', '_')
    title += ('_' + timezone.now().strftime("%Y-%m-%d_%H%M"))
    if not title.endswith(ends_with):
        title += ends_with
    return title


def _can_change_or_view(model, user):
    model_name = model._meta.model_name
    app_label = model._meta.app_label
    can_change = user.has_perm(app_label + '.change_' + model_name)
    can_view = user.has_perm(app_label + '.view_' + model_name)

    return can_change or can_view


def report_to_list(queryset, display_fields, user):
    model_class = queryset.model
    objects = queryset
    message = ""

    if not _can_change_or_view(model_class, user):
        return [], 'Permission Denied'

    new_display_fields = []

    for display_field in display_fields:
        field_list = display_field.split('__')
        field = field_list[-1]
        path = '__'.join(field_list[:-1])

        if path:
            path += '__'

        df = DisplayField(path, field)
        new_display_fields.append(df)

    display_fields = new_display_fields
    display_field_paths = []

    for i, display_field in enumerate(display_fields):
        model = get_model_from_path_string(model_class, display_field.path)

        if not model or _can_change_or_view(model, user):
            display_field_key = display_field.path + display_field.field

            display_field_paths.append(display_field_key)

        else:
            message += 'Error: Permission denied on access to {0}.'.format(
                display_field.name
            )

    values_list = objects.values_list(*display_field_paths)
    values_and_properties_list = [list(row) for row in values_list]

    return values_and_properties_list, message


def build_sheet(data, ws, sheet_name='report', header=None, widths=None):
    first_row = 1
    column_base = 1

    ws.title = re.sub(r'\W+', '', sheet_name)[:30]
    if header:
        for i, header_cell in enumerate(header):
            cell = ws.cell(row=first_row, column=i + column_base)
            cell.value = header_cell
            cell.font = Font(bold=True)
            if widths:
                ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

    for row in data:
        for i in range(len(row)):
            item = row[i]
            if isinstance(item, str):
                try:
                    row[i] = text_type(item)
                except UnicodeDecodeError:
                    row[i] = text_type(item.decode('utf-8', 'ignore'))
            elif isinstance(item, datetime.datetime):
                row[i] = item.replace(tzinfo=None)
            elif type(item) is dict:
                row[i] = text_type(item)
        try:
            ws.append(row)
        except ValueError as e:
            ws.append([e.message])
        except:
            ws.append(['Unknown Error'])


def list_to_workbook(data, title='report', header=None, widths=None):
    wb = Workbook()
    title = re.sub(r'\W+', '', title)[:30]

    if isinstance(data, dict):
        i = 0
        for sheet_name, sheet_data in data.items():
            if i > 0:
                wb.create_sheet()
            ws = wb.worksheets[i]
            build_sheet(
                sheet_data, ws, sheet_name=sheet_name, header=header)
            i += 1
    else:
        ws = wb.worksheets[0]
        build_sheet(data, ws, header=header, widths=widths)
    return wb


def build_xlsx_response(wb, title="report"):
    title = generate_filename(title, '.xlsx')
    myfile = BytesIO()
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        myfile.write(tmp.read())
    response = HttpResponse(
        myfile.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % title
    response['Content-Length'] = myfile.tell()
    return response


def list_to_xlsx_response(data, title='report', header=None,
                          widths=None):
    wb = list_to_workbook(data, title, header, widths)
    return build_xlsx_response(wb, title=title)


def list_to_csv_response(data, title='report', header=None, widths=None):
    response = HttpResponse(content_type="text/csv; charset=UTF-8")
    cw = csv.writer(response)
    for row in chain([header] if header else [], data):
        cw.writerow([force_text(s).encode(response.charset) for s in row])
    return response


def list_to_html_response(data, title='', header=None):
    html = render_to_string('export_action/report_html.html', locals())
    return HttpResponse(html)
