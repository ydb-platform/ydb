# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

from openpyxl.utils.cell import coordinate_to_tuple
from six import PY3

from efc.interfaces.base import BaseExcelInterface, CellInfo
from efc.interfaces.errors import NamedRangeNotFound
from efc.rpn_builder.parser.operands import CellAddress
from efc.utils import datetime_to_openxml, parse_date


class OpenpyxlInterface(BaseExcelInterface):
    def __init__(self, wb, *args, **kwargs):
        self.wb = wb
        super(OpenpyxlInterface, self).__init__(*args, **kwargs)

    def _serialize_value(self, value, data_type):
        if data_type == 'd':
            value = datetime_to_openxml(value)
        return value

    def _deserialize_value(self, value, data_type):
        if data_type == 'd':
            value = parse_date(value)
        return value

    def get_cell_formula_hyperlink(self, cell_index, ws_name):
        row, column = coordinate_to_tuple(cell_index)
        return self._get_cell_formula_hyperlink(CellAddress(ws_name, row, column, False, False))

    def calc_cell(self, cell_index, ws_name):
        """
        Calculate the cell formula by str index.
        If the cell does not have the formula it returns the cell's value.
        :type cell_index: basestring
        :type ws_name: basestring
        """
        row, column = coordinate_to_tuple(cell_index)
        value, _ = self._cell_to_value(CellAddress(ws_name, row, column, False, False))
        cell = self.wb[ws_name]._get_cell(row, column)
        return self._deserialize_value(value, cell.data_type)

    def _get_cell_info(self, address):
        cell = self.wb[address.ws_name]._get_cell(address.row, address.column)
        value = cell.value
        if cell.data_type != 'f':
            return CellInfo(self._serialize_value(value, cell.data_type))
        else:
            return CellInfo(None, cell.value[1:])

    if PY3:
        def _get_named_range_formula(self, name, ws_name):
            check = []
            if ws_name is not None:
                check.append(self.wb[ws_name])
            check.append(self.wb)
            for obj in check:
                if name in obj.defined_names:
                    return obj.defined_names[name].attr_text
            else:
                raise NamedRangeNotFound
    else:
        def _get_named_range_formula(self, name, ws_name):
            local_sheet_id = None
            if ws_name is not None:
                local_sheet_id = self.wb.sheetnames.index(ws_name)

            result = None
            for named_range in self.wb.defined_names.definedName:
                if named_range.name == name:
                    result = named_range.attr_text
                    if named_range.localSheetId == local_sheet_id:
                        break

            if result is None:
                raise NamedRangeNotFound
            else:
                return result

    def _max_row(self, ws_name):
        return self.wb[ws_name].max_row

    def _min_row(self, ws_name):
        return self.wb[ws_name].min_row

    def _max_column(self, ws_name):
        return self.wb[ws_name].max_column

    def _min_column(self, ws_name):
        return self.wb[ws_name].min_column

    def _has_worksheet(self, ws_name):
        return ws_name in self.wb.sheetnames
